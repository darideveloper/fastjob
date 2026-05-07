import logging
import traceback
from datetime import timedelta

import openpyxl
from celery import shared_task
from django.conf import settings
from django.db import transaction
from django.utils import timezone

from .models import CompanyImportBatch
from .importers import import_companies_from_xlsx
from .queries import bust_filter_caches

logger = logging.getLogger(__name__)


# Defensive ceiling against pathological xlsx files where ws.max_row reports
# 1_048_576 (Excel's row limit) due to phantom whole-column formatting.
_MAX_ROWS_CLAMP = 5_000_000


def _preflight_total_rows(file_path):
    """Open the workbook in read-only/streaming mode and return the number of
    data rows (excluding the header), clamped to a defensive ceiling.

    Tries the cheap path first (`ws.max_row`, which reads the worksheet's
    `<dimension>` XML element). When the dimension element is missing
    (`max_row is None`) or appears to be a placeholder (`<= 1`) — common for
    xlsx files exported by non-Excel tools — calls `ws.reset_dimensions()` to
    clear the (potentially wrong) bound and falls back to streaming through
    the rows. The reset is critical: openpyxl's `iter_rows` internally honors
    `self.max_row` as an upper bound, so a placeholder `<dimension>A1:B1</...>`
    would otherwise stop the iteration at row 1 and the fallback would also
    return 0. In `read_only=True` mode the streaming walk has bounded
    memory, so even ~150k rows take ~1-2s.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        raw = ws.max_row
        raw = int(raw) if raw is not None else 0
        if raw <= 1:
            ws.reset_dimensions()
            raw = sum(1 for _ in ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return min(max(raw - 1, 0), _MAX_ROWS_CLAMP)


@shared_task
def process_company_import(batch_id):
    try:
        batch = CompanyImportBatch.objects.get(id=batch_id)
    except CompanyImportBatch.DoesNotExist:
        return

    batch.status = "PROCESSING"
    batch.save(update_fields=["status", "updated_at"])

    try:
        batch.total_rows = _preflight_total_rows(batch.file.path)
        batch.save(update_fields=["total_rows", "updated_at"])

        created, updated, errors, blacklisted_skipped = import_companies_from_xlsx(
            batch.file.path,
            batch=batch,
            chunk_size=settings.COMPANY_IMPORT_CHUNK_SIZE,
        )

        try:
            batch.file.delete(save=False)
        except FileNotFoundError:
            pass
        batch.file.name = ""

        batch.created_count = created
        batch.updated_count = updated
        batch.blacklisted_skipped = blacklisted_skipped
        batch.error_log = errors
        batch.status = "COMPLETED"
        # Reconcile the denominator: the preflight ws.max_row estimate counts
        # trailing blank rows; pinning total_rows to processed_rows on success
        # makes the dashboard land on exactly 100 %.
        batch.total_rows = batch.processed_rows
        batch.save(
            update_fields=[
                "file",
                "status",
                "total_rows",
                "created_count",
                "updated_count",
                "blacklisted_skipped",
                "error_log",
                "updated_at",
            ]
        )

        # Bust the filter caches exactly once per successful import. The call
        # lives here (not in the importer) because the importer commits in
        # multiple per-chunk transactions; transaction.on_commit only fires
        # on the outermost block, and we want exactly one bust.
        transaction.on_commit(bust_filter_caches)

    except Exception as e:
        batch.status = "FAILED"
        # Preserve any progress committed by completed chunks; only append
        # the failure entries to error_log and flip status.
        existing_errors = list(batch.error_log or [])
        existing_errors.extend(
            [
                f"Error del sistema: {str(e)}",
                traceback.format_exc(),
                {"phase": "process", "file_path": batch.file.name},
            ]
        )
        batch.error_log = existing_errors
        batch.save(update_fields=["status", "error_log", "updated_at"])


@shared_task
def purge_stale_company_import_files():
    cutoff = timezone.now() - timedelta(days=settings.COMPANY_IMPORT_FILE_RETENTION_DAYS)
    stale = CompanyImportBatch.objects.filter(created_at__lt=cutoff).exclude(file="")

    deleted = 0
    missing = 0
    errored = 0

    for batch in stale:
        try:
            if batch.file.storage.exists(batch.file.name):
                batch.file.delete(save=False)
                deleted += 1
            else:
                missing += 1
            batch.file.name = ""
            batch.save(update_fields=["file", "updated_at"])
        except Exception as exc:
            logger.error("purge_stale_company_import_files: error purging batch %s: %s", batch.id, exc)
            errored += 1

    logger.info(
        "purge_stale_company_import_files: deleted=%d missing=%d errored=%d",
        deleted,
        missing,
        errored,
    )
