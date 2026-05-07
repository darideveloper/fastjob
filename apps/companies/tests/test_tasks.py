import os
from datetime import timedelta
from unittest.mock import patch

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.utils import timezone

from apps.companies.models import Company, CompanyImportBatch
from apps.companies.tasks import process_company_import, purge_stale_company_import_files
from apps.companies.tests.test_importers import make_xlsx


def _make_batch_with_local_file(tmp_path, filename="test.xlsx", content=None):
    path = make_xlsx(
        ["empresa", "email", "actividad", "poblacion"],
        [["Acme", "hr@acme.com", "Tech", "Madrid"]],
    )
    with open(path, "rb") as f:
        content = content or f.read()
    os.remove(path)

    uploaded_file = SimpleUploadedFile(filename, content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    batch = CompanyImportBatch.objects.create(status="PENDING")
    batch.file.save(filename, uploaded_file, save=True)
    return batch


@pytest.mark.django_db
def test_process_company_import_task_success(tmp_path):
    with override_settings(COMPANY_IMPORT_LOCAL_PATH=str(tmp_path)):
        batch = _make_batch_with_local_file(tmp_path)
        file_name = batch.file.name

        process_company_import(batch.id)

        batch.refresh_from_db()
        assert batch.status == "COMPLETED"
        assert batch.created_count == 1
        assert batch.updated_count == 0
        assert batch.error_log == []
        assert Company.objects.count() == 1
        # File should be deleted on success
        assert not batch.file.name


@pytest.mark.django_db
def test_process_company_import_task_handles_invalid_file(tmp_path):
    with override_settings(COMPANY_IMPORT_LOCAL_PATH=str(tmp_path)):
        upload = SimpleUploadedFile("test.xlsx", b"not a real excel file", content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        batch = CompanyImportBatch.objects.create(status="PENDING")
        batch.file.save("test.xlsx", upload, save=True)
        file_abs_path = batch.file.path

        process_company_import(batch.id)

        batch.refresh_from_db()
        assert batch.status == "FAILED"
        assert batch.created_count == 0
        assert batch.error_log
        assert "Error del sistema" in batch.error_log[0]
        # File should be kept on failure
        assert os.path.exists(file_abs_path)


@pytest.mark.django_db
def test_successful_processing_deletes_local_file(tmp_path):
    with override_settings(COMPANY_IMPORT_LOCAL_PATH=str(tmp_path)):
        batch = _make_batch_with_local_file(tmp_path)
        file_abs_path = batch.file.path

        process_company_import(batch.id)

        batch.refresh_from_db()
        assert batch.status == "COMPLETED"
        assert not os.path.exists(file_abs_path)


@pytest.mark.django_db
def test_failed_processing_leaves_file_on_disk(tmp_path):
    with override_settings(COMPANY_IMPORT_LOCAL_PATH=str(tmp_path)):
        upload = SimpleUploadedFile("test.xlsx", b"not a real excel file")
        batch = CompanyImportBatch.objects.create(status="PENDING")
        batch.file.save("test.xlsx", upload, save=True)
        file_abs_path = batch.file.path

        process_company_import(batch.id)

        batch.refresh_from_db()
        assert batch.status == "FAILED"
        assert os.path.exists(file_abs_path)


@pytest.mark.django_db
def test_process_company_import_writes_total_rows_after_preflight(tmp_path):
    with override_settings(COMPANY_IMPORT_LOCAL_PATH=str(tmp_path)):
        path = make_xlsx(
            ["empresa", "email"],
            [
                ["A", "a@x.com"],
                ["B", "b@x.com"],
                ["C", "c@x.com"],
                ["D", "d@x.com"],
                ["E", "e@x.com"],
            ],
        )
        with open(path, "rb") as f:
            content = f.read()
        os.remove(path)

        upload = SimpleUploadedFile("test.xlsx", content)
        batch = CompanyImportBatch.objects.create(status="PENDING")
        batch.file.save("test.xlsx", upload, save=True)

        process_company_import(batch.id)

        batch.refresh_from_db()
        assert batch.status == "COMPLETED"
        assert batch.total_rows == 5
        assert batch.processed_rows == 5
        assert batch.processed_rows == batch.total_rows


def test_preflight_falls_back_to_streaming_when_max_row_is_unreliable():
    """Regression: the user reported `total_rows = 0` after upload because
    their xlsx producer wrote the worksheet without a trustworthy
    `<dimension>` element, so openpyxl's `ws.max_row` returned `None` (no
    dimension) or a placeholder `1` (header-only), and the preflight
    computed `max((None or 1) - 1, 0) = 0`.

    The fix detects the unreliable bound, calls `ws.reset_dimensions()` to
    clear openpyxl's internal `_max_row` cap (otherwise `iter_rows` would
    also honor the bogus bound), and stream-counts the rows.

    Faithful simulation: openpyxl populates `_max_row` from the dimension
    element during worksheet construction. We patch `__init__` to overwrite
    it post-init, which is exactly the state a non-conformant producer
    creates.
    """
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet
    from apps.companies.tasks import _preflight_total_rows

    path = make_xlsx(
        ["empresa", "email"],
        [
            ["A", "a@x.com"],
            ["B", "b@x.com"],
            ["C", "c@x.com"],
            ["D", "d@x.com"],
            ["E", "e@x.com"],
        ],
    )

    original_init = ReadOnlyWorksheet.__init__

    def init_with_max_row(forced_max_row):
        def _init(self, *args, **kwargs):
            original_init(self, *args, **kwargs)
            self._max_row = forced_max_row
        return _init

    try:
        # Case 1: producer omitted <dimension> entirely (max_row = None).
        with patch.object(
            ReadOnlyWorksheet, "__init__", init_with_max_row(None)
        ):
            assert _preflight_total_rows(path) == 5

        # Case 2: producer wrote a placeholder dimension (max_row = 1).
        # reset_dimensions() must clear it before iter_rows runs.
        with patch.object(
            ReadOnlyWorksheet, "__init__", init_with_max_row(1)
        ):
            assert _preflight_total_rows(path) == 5
    finally:
        os.remove(path)


@pytest.mark.django_db
def test_process_company_import_schedules_filter_cache_bust_once(tmp_path):
    with override_settings(COMPANY_IMPORT_LOCAL_PATH=str(tmp_path)):
        batch = _make_batch_with_local_file(tmp_path)

        # pytest-django's @django_db wraps each test in a transaction that
        # is rolled back, so transaction.on_commit callbacks never actually
        # fire. We instead assert the call was *scheduled* exactly once,
        # with bust_filter_caches as the callback.
        with patch("apps.companies.tasks.transaction.on_commit") as mock_oc, \
             patch("apps.companies.tasks.bust_filter_caches") as mock_bust:
            process_company_import(batch.id)

        assert mock_oc.call_count == 1
        assert mock_oc.call_args[0][0] is mock_bust


@pytest.mark.django_db
def test_purge_removes_only_stale_files(tmp_path):
    with override_settings(
        COMPANY_IMPORT_LOCAL_PATH=str(tmp_path),
        COMPANY_IMPORT_FILE_RETENTION_DAYS=7,
    ):
        upload_old = SimpleUploadedFile("old.xlsx", b"data")
        old_batch = CompanyImportBatch.objects.create(status="FAILED")
        old_batch.file.save("old.xlsx", upload_old, save=True)
        old_file_path = old_batch.file.path
        # Back-date to exceed retention window
        CompanyImportBatch.objects.filter(id=old_batch.id).update(
            created_at=timezone.now() - timedelta(days=8)
        )

        upload_recent = SimpleUploadedFile("recent.xlsx", b"data")
        recent_batch = CompanyImportBatch.objects.create(status="FAILED")
        recent_batch.file.save("recent.xlsx", upload_recent, save=True)
        recent_file_path = recent_batch.file.path

        purge_stale_company_import_files()

        assert not os.path.exists(old_file_path)
        old_batch.refresh_from_db()
        assert not old_batch.file.name

        assert os.path.exists(recent_file_path)
        recent_batch.refresh_from_db()
        assert recent_batch.file.name
